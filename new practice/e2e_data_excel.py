import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

#excel file path

excel_file=r"D:\STUDY AUTO\Selenium with python\OrangeHRM.xlsx"
sheet_name="Sheet1"

#load workbook
workbook=load_workbook(excel_file)
sheet=workbook[sheet_name]

#chorme iinvoke

driver=webdriver.Chrome()
driver.maximize_window()

#loop through the excel

for row in range(2,sheet.max_row+1):
    username=sheet.cell(row=row,column=1).value
    password=sheet.cell(row=row,column=2).value
    #varOcg="Row_" + str(row)
    result_col=3            #column sresult

    try:
        driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
        time.sleep(3)

        #enter username pwd
        driver.find_element(By.NAME,"username").send_keys(username)
        driver.find_element(By.NAME,"password").send_keys(password)
        driver.find_element(By.XPATH,"//button[@type='submit']").click()
        time.sleep(5)

        # title=driver.title
        # #write result in excel
        # sheet.cell(row=row,column=result_col).value=title
        # print(f"{varOcg}: Login success -> Title written in Excel")

        # Check if Dashboard is displayed
        if "dashboard" in driver.current_url.lower():
            sheet.cell(row=row,column=result_col).value="Login Success"
            print(f" Row {row}: Login Successful ✅")

            #logout
            driver.find_element(By.XPATH,"//p[@class='oxd-userdropdown-name']").click()
            time.sleep(1)
            driver.find_element(By.XPATH,"//a[text()='Logout']").click()
            time.sleep(2)

        else:
            sheet.cell(row=row,column=result_col).value="Login Failed"
            print(f" Row {row}: Login Failed ❌")

    except Exception as e:

        print(f" Row {row}: Login failed -> {str(e)}")
        sheet.cell(row=row,column=result_col).value="Error Message"

#save workbook
workbook.save(excel_file)
print("Data saved into excel")
driver.quit()
















