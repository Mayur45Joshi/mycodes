import openpyxl

# file="D:\STUDY AUTO\Excel.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook["Sheet1"]
# rows=sheet.max_row    #no of rows
# cols=sheet.max_column   #no of columns

#reading all columns and rows

# for r in range(1,rows+1):
#     for c in range(1,cols+1):
#         print(sheet.cell(r,c).value)
#     print()


####################################################

#writing data to excel

# file="D:\STUDY AUTO\Excel.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook.active
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Welcome"
# workbook.save(file)



#################################################


#different data entere

file="D:\STUDY AUTO\Excel.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active

sheet.cell(1,1).value=123
sheet.cell(1,2).value="Mj"
sheet.cell(1,3).value="Engineer"

sheet.cell(2,1).value=124
sheet.cell(2,2).value="Rohan"
sheet.cell(2,3).value="Plumber"

sheet.cell(3,1).value=125
sheet.cell(3,2).value="Sita"
sheet.cell(3,3).value="Girl"

workbook.save(file)