# from openpyxl import Workbook
#
# # Create workbook
# wb = Workbook()
#
# # Sheet 1
# ws1 = wb.active
# ws1.title = "Students"
# ws1.append(["ID", "Name", "Marks"])
# ws1.append([1, "Ravi", 85])
# ws1.append([2, "Anita", 90])
#
# # Sheet 2
# ws2 = wb.create_sheet(title="Projects")
# ws2.append(["ProjectID", "ProjectName", "Status"])
# ws2.append([101, "API Testing", "Completed"])
# ws2.append([102, "UI Automation", "In Progress"])
#
# # Save file
# wb.save(r"D:\STUDY AUTO\employees.xlsx")
#
# print("Data written in multiple sheets successfully!")

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>.

# from openpyxl import Workbook
#
# # Create workbook
# wb = Workbook()
# ws = wb.active
# ws.title = "Employees"
#
# # Header row
# ws.append(["ID", "Name", "Role", "Salary"])
#
# # Data stored in list of lists
# employees = [
#     [1, "Mayur", "QA Engineer", 50000],
#     [2, "Priya", "Developer", 60000],
#     [3, "Amit", "Manager", 75000],
#     [4, "Sita", "Tester", 45000],
# ]
#
# # Write data row by row using loop
# for emp in employees:
#     ws.append(emp)
#
# # Save file
# wb.save(r"D:\STUDY AUTO\employees.xlsx")
#
# print("Data written successfully!")
# # #writting data from excel
# #
# # for row in ws.iter_rows(values_only=True):
# #     print(row)
#
# #Reading Cell by Cell
# # Print specific cell values
# print("Employee Name in row 2:", ws["B2"].value)
# print("Role in row 3:", ws["C3"].value)
#
# # Print all rows except header
# for row in ws.iter_rows(min_row=1, values_only=True):
#     emp_id, name, role, salary = row
#     print(f"ID={emp_id}, Name={name}, Role={role}, Salary={salary}")


#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

# #taking Input from User with Loop
#
# from openpyxl import Workbook
#
# wb = Workbook()
# ws = wb.active
# ws.title = "Employees"
#
# ws.append(["ID", "Name", "Role", "Salary"])
#
# # Number of records
# n = int(input("Enter number of employees: "))
#
# for i in range(n):
#     emp_id = int(input("Enter Employee ID: "))
#     name = input("Enter Employee Name: ")
#     role = input("Enter Role: ")
#     salary = int(input("Enter Salary: "))
#
#     ws.append([emp_id, name, role, salary])
#
# wb.save(r"D:\STUDY AUTO\employees.xlsx")
# print("Data written successfully!")
#
# # #writting data from excel
# #
# # for row in ws.iter_rows(values_only=True):
# #     print(row)
#
# #Reading Cell by Cell
# # Print specific cell values
# print("Employee Name in row 2:", ws["B2"].value)
# print("Role in row 3:", ws["C3"].value)
#
# # Print all rows except header
# for row in ws.iter_rows(min_row=1, values_only=True):
#     emp_id, name, role, salary = row
#     print(f"ID={emp_id}, Name={name}, Role={role}, Salary={salary}")

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>.

# #all in one adding and reading
#
# from openpyxl import Workbook, load_workbook
# # -------------------------------
# # STEP 1: Write data to Excel
# # -------------------------------
# wb = Workbook()
# ws = wb.active
# ws.title = "Employees"
#
# # Add header row
# ws.append(["ID", "Name", "Role", "Salary"])
#
# # Employee data
# employees = [
#     [1, "Mayur", "QA Engineer", 50000],
#     [2, "Priya", "Developer", 60000],
#     [3, "Amit", "Manager", 75000],
#     [4, "Sita", "Tester", 45000],
# ]
# # Write all rows
# for emp in employees:
#     ws.append(emp)
#
# # Save file
# file_name = r"D:\STUDY AUTO\employees.xlsx"
# wb.save(file_name)
# print(f"✅ Data written successfully to {file_name}")
# # -------------------------------
# # STEP 2: Read data back from Excel
# # -------------------------------
# wb2 = load_workbook(file_name)
# ws2 = wb2.active
#
# print("\n📖 Reading data from Excel:\n")
#
# # Print each row
# for row in ws2.iter_rows(values_only=True):
#     print(row)

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>..

# #writting data into same excel sheet
#
# from openpyxl import load_workbook
#
# # File to update
# file_name = r"D:\STUDY AUTO\employees.xlsx"
#
# # Load the existing workbook
# wb = load_workbook(file_name)
# ws = wb.active
#
# # New employee data to add
# new_employees = [
#     [5, "Rohit", "DevOps Engineer", 65000],
#     [6, "Anjali", "Business Analyst", 55000],
# ]
#
# # Append new rows
# for emp in new_employees:
#     ws.append(emp)
#
# # Save changes
# wb.save(file_name)
#
# print(f"✅ New rows added successfully to {file_name}")

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>..

from openpyxl import Workbook,load_workbook

#souce_file
src_f=r"D:\STUDY AUTO\employees.xlsx"
src_wb=load_workbook(src_f)
src_ws=src_wb.active

#destination file new workbook
des_f=r"D:\STUDY AUTO\employees1.xlsx"
des_wb=Workbook()
des_ws=des_wb.active
des_ws.title="CopyData"

for row in src_ws.iter_rows(values_only=True):
    des_ws.append(row)

des_wb.save(des_f)
print("Data Copied")